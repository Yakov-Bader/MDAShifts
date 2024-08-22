from collections import defaultdict
from typing import List
from openpyxl import Workbook
import datetime, calendar
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from shifts_algo.volunteer import Volunteer
from shifts_algo.shift import Shift


class MDASheet:

    def __init__(self):
        self.workbook = Workbook()
        self.sheet = self.workbook.active

        # set general styles
        self.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        self.thin_border_style = Side(border_style="thin", color="000000")
        self.thick_border_style = Side(border_style="thick", color="000000")
        self.border = Border(
            bottom=self.thin_border_style,
            top=self.thin_border_style,
            left=self.thin_border_style,
            right=self.thin_border_style,
        )
        self.border_without_right = Border(
            bottom=self.thin_border_style,
            top=self.thin_border_style,
            left=self.thin_border_style,
            right=self.thick_border_style,
        )

    def create_headers(self) -> None:
        """
        This function creates the headers for the table, and also colors in the columns
        """
        
        # there are two rows of main headers
        headers = [
            [
                "נהגים חיליק עזרזר 052-3462562",
                "יום א' אריק אלישע 054-6896101",
                "יום ג' ירון בר זכאי  054-6303308",
                "יום ה' אלכסיי פורטנוב 052-3972002",
                "לוח משמרות מתנדבים- אוגוסט 2024",
                'מקרא צבעים: נהגים/נהגים משתלמים  בוגרים  חובשים משתלמים  חונכים  מע"רים  חניכים',
            ],
            [
                'אט"ן בקי גילאור-0542009548',
                "יום ב'- דני זרק 058-6802118",
                "יום ד' נדין כהן 050-4084069",
                "יום ו'+שבת מיכל מאיה סלע 052-3132598",
                "תחנת מודיעין - מרחב איילון",
            ],
        ]
        # Create general headers
        for j, header in enumerate(headers):
            i = 4
            for val in header:
                end = i + 5 if i > 18 else i + 2
                self._merge_and_set_value(
                    start_row=j + 1,
                    start_col=i,
                    end_row=j + 1,
                    end_col=end,
                    value=val,
                )
                i = end + 1

        # here are added the rows for the shift headers
        shift_names = [
            {"name": "בוקר תגבור", "color": "339966"},
            {"name": 'אט"ן', "color": "ff6600"},
            {"name": 'אט"ן', "color": "ff6600"},
            {"name": "רגיל - תקן (ק.ס.)", "color": "bfbfbf"},
            {"name": "בוקר תחנה | ערב רגיל - מתנדבים", "color": "339966"},
            {"name": "ערב רגיל - מתנדבים", "color": "43c383"},
            {"name": "ערב רגיל - מתנדבים", "color": "339966"},
            {"name": "ערב רגיל - מתנדבים", "color": "43c383"},
            {"name": "ערב רגיל - מתנדבים", "color": "339966"},
            {"name": "ערב רגיל - מתנדבים", "color": "43c383"},
            {"name": "ערב רגיל - מתנדבים", "color": "339966"},
        ]

        # Create all shifts columns
        column_location: int = 4
        for shift in shift_names:
            col_end: int = (
                column_location + 3
                if len(shift["name"].strip()) > 5
                else column_location + 1
            )
            # shift name
            self._merge_and_set_value(
                start_row=3,
                start_col=column_location,
                end_row=3,
                end_col=col_end,
                value=shift["name"],
            )

            color_fill = PatternFill(
                fill_type="darkUp",
                start_color=shift["color"],
                end_color=shift["color"],
            )

            self.sheet[self._cell_coordinates_to_reference(3, column_location)].fill = (
                color_fill
            )

            # Fill columns according to the shift
            for column in range(column_location, col_end + 1):
                for row in self.sheet.iter_rows(
                    min_col=column, max_col=column, min_row=4, max_row=94
                ):
                    for cell in row:
                        cell.fill = color_fill

            # border edge of shift
            border = Border(right=self.thick_border_style)
            for row in self.sheet.iter_rows(
                min_col=col_end, max_col=col_end, min_row=3, max_row=94
            ):
                for cell in row:
                    cell.border = border

            # define the volunteer column in the shift
            for row in self.sheet.iter_rows(
                min_col=column_location, max_col=col_end, min_row=4, max_row=4
            ):
                for i, cell in enumerate(row):
                    cell.value = (
                        "נהג" if len(shift["name"].strip()) > 5 and not i else "מתנדב"
                    )

            column_location = col_end + 1

    def create_sides(self, next_month_year: int, next_month_month:int) -> None:
        """
        Create the row definitions, and give color for weekends
        """
        num_days: int = calendar.monthrange(next_month_year, next_month_month)[1]
        days = [datetime.date(next_month_year, next_month_month, day) for day in range(1, num_days + 1)]

        # Give values to date headers
        for i, val in enumerate(["יום", "תאריך", "שעה"]):
            self._merge_and_set_value(
                start_row=3,
                start_col=i + 1,
                end_row=4,
                end_col=i + 1,
                value=val,
            )

        saturday_fill = PatternFill(
            fill_type="solid", start_color="00b0f0", end_color="000000FF"
        )

        # For each day, create three shifts
        for l in range(len(days) * 3):
            english_day = str(days[int(l / 3)].strftime("%A"))

            val = ""
            if l % 3 == 0:
                val = "לילה"
                # Create date cell
                self._merge_and_set_value(
                    start_row=l + 5,
                    start_col=2,
                    end_row=l + 7,
                    end_col=2,
                    value=str(days[int(l / 3)].strftime("%Y/%m/%d")),
                )

                # Create day cell
                self._merge_and_set_value(
                    start_row=l + 5,
                    start_col=1,
                    end_row=l + 7,
                    end_col=1,
                    value=english_day,
                )

                # Color date and day if it is Saturday
                if english_day == "Saturday":
                    self.sheet[self._cell_coordinates_to_reference(l + 5, 1)].fill = (
                        saturday_fill
                    )
                    self.sheet[self._cell_coordinates_to_reference(l + 5, 2)].fill = (
                        saturday_fill
                    )
            elif l % 3 == 1:
                val = "בוקר"
            else:
                val = "ערב"

            # create shift time
            self.sheet.cell(row=l + 5, column=3, value=val)

            # Color row if it is Saturday
            if english_day == "Saturday" or (english_day == "Friday" and val == "ערב"):

                for row in self.sheet.iter_rows(
                    min_col=3, min_row=l + 5, max_row=l + 5
                ):
                    for cell in row:
                        cell.fill = saturday_fill

    def fill_table(self, shifts: List[Shift]) -> None:
        """this function adds the shifts to the excel, and at the end adds the unassigned volunteers to the table

        Args:
            shifts (List[Shift]): a list of the assigned shifts with volunteers
        """
        for shift in shifts:
            # define what column to start
            row_num = 2 + shift.shift_date
            col_start = 13
            if "תגבור" in shift.shift_name:
                col_start = 5
            elif 'אט"ן' in shift.shift_name:
                col_start = 8
            else:
                while (
                    self.sheet[
                        self._cell_coordinates_to_reference(row_num, col_start)
                    ].value
                    is not None
                ):
                    col_start += 4
                if shift.gets_driver:
                    col_start -= 1
            
            # give the cells a value
            for i in range(col_start, col_start + len(shift.volunteers)):
                volunteer = shift.volunteers[i - col_start]
                cell = self.sheet[self._cell_coordinates_to_reference(row_num, i)]
                cell.value = f"{volunteer.name}({volunteer.get_assigned_shifts_amount()}/{volunteer.get_wanted_amount()})"
                font = Font(color=self._get_volunteer_color(volunteer.rank))
                cell.font = font
        
    
    def add_unsigned_to_table(self, shifts: List[Shift]) -> None:
        """
        add unassigned volunteers to end of table
        Args:
            shifts (List[Shift]): shift list after users where assigned
        """
        # create a dict, that puts in the same list shifts that are the same
        shifts_dict = defaultdict(list)
        for shift in shifts:
            shifts_dict[f'{"".join(shift.shift_name.split())}{shift.shift_date}'].append(shift)

        # merge shifts that are the same
        merged_shifts: List[Shift] = []
        for date_name, shifts in shifts_dict.items():
            if len(shifts) > 1:
                merged_shift = Shift(shifts[0].shift_date, shifts[0].shift_name, shifts[0].get_max_volunteers())
                for shift in shifts:
                    merged_shift.add_volunteers(shift.get_volunteers())
                    merged_shift.add_wanters(shift.get_ordered_unsigned_volunteers())
                merged_shifts.append(merged_shift)
            else:
                merged_shifts.append(shifts[0])
        
        # present the unassigned of each shift            
        for shift in merged_shifts:
            row_num: int = 2+ shift.shift_date
            unassigned_volunteers: List[Volunteer] = (
                shift.get_ordered_unsigned_volunteers()
            )
            if len(unassigned_volunteers):
                # find the begging column
                col_start: int = 44
                while (
                    self.sheet[
                        self._cell_coordinates_to_reference(row_num, col_start)
                    ].value
                    is not None
                ):
                    col_start += 1
                cell = self.sheet[
                    self._cell_coordinates_to_reference(row_num, col_start)
                ]
                
                # add the shift name with the unassigned volunteers
                cell.value = shift.shift_name 
                col_start += 1
                
                # add the unassigned volunteers to the table
                for i, volunteer in enumerate(unassigned_volunteers):
                    cell = self.sheet[
                        self._cell_coordinates_to_reference(row_num, i + col_start)
                    ]
                    cell.value = f"{volunteer.name}({volunteer.get_assigned_shifts_amount()}/{volunteer.get_wanted_amount()})"
                    font = Font(color=self._get_volunteer_color(volunteer.rank))
                    color_fill = PatternFill(
                        fill_type="darkUp",
                        start_color="43c383",
                        end_color="43c383",
                    )
                    cell.fill = color_fill
                    cell.font = font

    def display(self) -> None:
        """
        Here the table is displayed into the workbook
        """
        # Align all cell, and give them border, wouthout removing old border
        for row in self.sheet.iter_rows(min_row=1, min_col=1):
            for cell in row:
                cell.alignment = self.alignment
                if cell.border.right and cell.border.right.style is None:
                    cell.border = self.border
                else:
                    cell.border = self.border_without_right

        # Give size to all cells
        for col in self.sheet.columns:
            column_letter = get_column_letter(col[0].column)
            self.sheet.column_dimensions[column_letter].width = 20

        for row in self.sheet.rows:
            self.sheet.row_dimensions[row[0].row].height = 17
        self.workbook.save("MDA_sheet.xlsx")

    def _merge_and_set_value(
        self, start_row: int, start_col: int, end_row: int, end_col: int, value: str
    ) -> None:
        """as the name declares, the function merges a cell and gives it a value 

        Args:
            start_row (int): begging row of cell to be merged
            start_col (int): begging column of cell to be merged
            end_row (int): end row of cell to be merged
            end_col (int): end column of cell to be merged
            value (str): _description_
        """
        # Merge the cells
        self.sheet.merge_cells(
            start_row=start_row,
            start_column=start_col,
            end_row=end_row,
            end_column=end_col,
        )
        # Set the value in the top-left cell of the merged range
        self.sheet.cell(row=start_row, column=start_col, value=value)

    def _cell_coordinates_to_reference(self, row: str, col: int) -> str:
        """Convert row and column numbers to Excel cell reference (e.g., (1, 1) -> 'A1')."""
        return get_column_letter(col) + str(row)

    def _get_volunteer_color(self, volunteer_rank: str) -> str:
        """
        For volunteer rank, get text color
        Args:
            volunteer_rank (str): volunteer MDA rank

        Returns:
            str: color representing volunteer rank
        """
        match volunteer_rank:
            case "נהג":
                return "0000ff"
            case "נהג משתלם":
                return "0000ff"
            case "חובש":
                return "ffff00"
            case "חובש משתלם":
                return "ff0000"
            case "חונך":
                return "385623"
            case 'מע"ר':
                return "ffffff"
            case "חניך":
                return "fb37d1"
            case _:
                return "00000000"
